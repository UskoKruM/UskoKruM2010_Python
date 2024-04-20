import openpyxl
from tabulate import tabulate

excel_dataframe = openpyxl.load_workbook("people.xlsx")

dataframe = excel_dataframe.active

data = []

for row in range(1, dataframe.max_row):
    _row = [row,]

    for col in dataframe.iter_cols(1, dataframe.max_column):
        _row.append(col[row].value)

    data.append(_row)

headers = ["#", "Id", "Name", "Age", "Company", "Email", "MAC Address"]
headers_align = (('center',) * 7)

print(tabulate(data, headers, tablefmt='fancy_grid', colalign=headers_align))
